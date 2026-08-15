"""
问题3：带时间要求的多日直升机排班（Python 3.11）。

方法：
1. 复用问题2的单架次空间取送、动态容量、燃油与加油可行性；
2. 保留个人级时间窗，生成带可行起飞时段的候选架次；
3. 用 CP-SAT 同时选择候选架次、分配24架具体飞机并安排起飞时刻；
4. 用时间相关大邻域不断扩充路线池；
5. 两阶段求解：先仅安排 mandatory 任务得到时间基准 T0，再在 T<=T0 下
   最大化 temporary 人数，随后严格优化其他指标。

重要边界：CP-SAT 可以证明当前已生成候选池内的最优性；候选路线不是全部潜在路线，
因此程序不声称整个原问题的全局最优。

默认输入：data/raw/distances.csv、data/raw/peopleQ3.csv
默认输出：docs/reference_formats/q3-routes.csv、q3-assignments.csv、q3-summary.json
依赖：python -m pip install ortools
"""

from __future__ import annotations

import argparse
import csv
import itertools
import json
import math
import os
import random
import sys
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib
import numpy as np
import pandas as pd

# 兼容旧版Matplotlib与新版NumPy组合。
if not hasattr(np, "row_stack"):
    np.row_stack = np.vstack

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from ortools.sat.python import cp_model
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "缺少 OR-Tools，请在当前 Python 3.11 环境执行：python -m pip install ortools"
    ) from exc

for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8", line_buffering=True)
        except (AttributeError, OSError):
            pass


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data" / "raw"
OUTPUT_DIR = BASE_DIR / "docs" / "reference_formats"
ROUTES_OUT = OUTPUT_DIR / "q3-routes.csv"
ASSIGN_OUT = OUTPUT_DIR / "q3-assignments.csv"
SUMMARY_OUT = OUTPUT_DIR / "q3-summary.json"
FIGURES_DIR = BASE_DIR / "outputs" / "figures"
TABLES_DIR = BASE_DIR / "outputs" / "tables"

HORIZON_START = datetime(2026, 8, 3, 0, 0)
PLANNING_DAYS = 7
DAY_MINUTES = 1440
TAKEOFF_OPEN = 6 * 60
TAKEOFF_CLOSE = 18 * 60
RETURN_CLOSE = 20 * 60
TURNAROUND = 30
HORIZON = PLANNING_DAYS * DAY_MINUTES

AIRPORTS = ("A01", "A02", "A03")
AIRPORT_SET = set(AIRPORTS)
LAND = "LAND"
GAS_STATIONS = ("F006", "F011", "F018", "F024", "F031", "F038", "F044", "F050")
GAS_SET = set(GAS_STATIONS)
AC = {
    "T1": {"seats": 12, "speed": 250.0, "burn": 3.4, "tank": 1000.0, "reserve": 150.0},
    "T2": {"seats": 16, "speed": 220.0, "burn": 2.5, "tank": 1150.0, "reserve": 150.0},
    "T3": {"seats": 19, "speed": 190.0, "burn": 2.9, "tank": 1600.0, "reserve": 200.0},
}
FLEET_COUNTS = {
    "A01": {"T1": 3, "T2": 3, "T3": 2},
    "A02": {"T1": 2, "T2": 4, "T3": 2},
    "A03": {"T1": 2, "T2": 3, "T3": 3},
}
TASK_ORDER = {"emergency": 0, "production": 1, "shift": 2, "temporary": 3}
MANDATORY_TYPES = {"emergency", "production", "shift"}
MAX_LANDINGS = 5
MAX_SEATS = 19
EPS = 1e-9
SCALE = 1000

D: Dict[str, Dict[str, float]] = {}


@dataclass(frozen=True)
class Person:
    index: int
    person_id: str
    origin: str
    destination: str
    earliest: int
    latest: int
    task_type: str

    @property
    def od(self) -> Tuple[str, str]:
        return self.origin, self.destination

    @property
    def mandatory(self) -> bool:
        return self.task_type in MANDATORY_TYPES


@dataclass(frozen=True)
class Route:
    airport: str
    aircraft_type: str
    stops: Tuple[str, ...]
    refuels: Tuple[bool, ...]
    pairs: Tuple[Tuple[str, str, int, int], ...]
    duration: int
    fuel_kg: float
    arrivals: Tuple[int, ...]
    departures: Tuple[int, ...]
    leg_distances: Tuple[float, ...]
    intransit_min: int
    pass_km: float
    avail_km: float

    @property
    def pair_map(self) -> Dict[Tuple[str, str], Tuple[int, int]]:
        return {(o, d): (p, q) for o, d, p, q in self.pairs}

    @property
    def signature(self) -> Tuple:
        return (
            self.airport, self.aircraft_type, self.stops,
            self.refuels, self.pairs,
        )


@dataclass
class Pattern:
    pattern_id: int
    people: Tuple[int, ...]
    route: Route
    start_intervals: Tuple[Tuple[int, int], ...]
    sources: set[str] = field(default_factory=set)

    @property
    def signature(self) -> Tuple:
        return self.people, self.route.signature, self.start_intervals


@dataclass(frozen=True)
class Selection:
    pattern_id: int
    aircraft_id: str
    start: int


@dataclass(frozen=True)
class Metrics:
    temporary_served: int = 0
    aircraft_time: int = 0
    person_intransit: int = 0
    sorties: int = 0
    fuel_kg: float = 0.0
    pass_km: float = 0.0
    avail_km: float = 0.0

    @property
    def seat_utilization(self) -> float:
        return 0.0 if self.avail_km <= EPS else self.pass_km / self.avail_km


@dataclass
class MasterResult:
    selections: List[Selection]
    metrics: Metrics
    stages: List[dict]
    pool_optimal: bool


def build_fleet() -> Tuple[Tuple[str, str, str], ...]:
    fleet = []
    for airport in AIRPORTS:
        for aircraft_type in AC:
            for number in range(1, FLEET_COUNTS[airport][aircraft_type] + 1):
                fleet.append((f"{airport}-{aircraft_type}-H{number:02d}", airport, aircraft_type))
    return tuple(fleet)


FLEET = build_fleet()
FLEET_BY_GROUP: Dict[Tuple[str, str], Tuple[str, ...]] = defaultdict(tuple)
for _aircraft_id, _airport, _aircraft_type in FLEET:
    FLEET_BY_GROUP[(_airport, _aircraft_type)] += (_aircraft_id,)
AIRCRAFT_INFO = {aid: (airport, ac_type) for aid, airport, ac_type in FLEET}


def parse_minute(raw: str) -> int:
    value = datetime.strptime(raw.strip(), "%Y-%m-%d %H:%M")
    delta = value - HORIZON_START
    minute = int(delta.total_seconds() // 60)
    if minute < 0 or minute > HORIZON:
        raise ValueError(f"时间超出规划区间：{raw}")
    return minute


def format_minute(minute: int) -> str:
    return (HORIZON_START + timedelta(minutes=int(minute))).strftime("%Y-%m-%d %H:%M")


def flight_min(distance: float, aircraft_type: str) -> int:
    return math.ceil(60.0 * distance / AC[aircraft_type]["speed"] - EPS)


def load_distances(path: Path) -> None:
    global D
    D = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        columns = [value.strip() for value in header[1:]]
        for row in reader:
            if not row:
                continue
            origin = row[0].strip()
            D[origin] = {
                destination: float(raw.strip())
                for destination, raw in zip(columns, row[1:])
            }
    if not AIRPORT_SET.issubset(D):
        raise ValueError("distances.csv 缺少 A01/A02/A03")
    for origin in D:
        if set(D) - set(D[origin]):
            raise ValueError(f"距离矩阵中 {origin} 行不完整")


def load_people(path: Path) -> List[Person]:
    people = []
    seen = set()
    required = {
        "person_id", "origin_id", "destination_id", "earliest_pickup_time",
        "latest_arrival_time", "task_type",
    }
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not required.issubset(reader.fieldnames or []):
            raise ValueError(f"peopleQ3.csv 必须包含 {sorted(required)}")
        for index, row in enumerate(reader):
            person_id = row["person_id"].strip()
            if not person_id or person_id in seen:
                raise ValueError(f"人员编号为空或重复：{person_id!r}")
            seen.add(person_id)
            origin = row["origin_id"].strip()
            destination = row["destination_id"].strip()
            task_type = row["task_type"].strip().lower()
            earliest = parse_minute(row["earliest_pickup_time"])
            latest = parse_minute(row["latest_arrival_time"])
            if task_type not in TASK_ORDER:
                raise ValueError(f"未知任务类型：{task_type}")
            if origin == destination or (origin in AIRPORT_SET | {LAND} and destination in AIRPORT_SET | {LAND}):
                raise ValueError(f"不支持的人员起终点：{person_id} {origin}->{destination}")
            if origin not in D and origin != LAND:
                raise ValueError(f"未知起点：{origin}")
            if destination not in D and destination != LAND:
                raise ValueError(f"未知终点：{destination}")
            if earliest > latest:
                raise ValueError(f"人员时间窗颠倒：{person_id}")
            people.append(Person(
                index, person_id, origin, destination,
                earliest, latest, task_type,
            ))
    if not people:
        raise ValueError("peopleQ3.csv 没有人员记录")
    return people


def canonical_people(ids: Iterable[int]) -> Tuple[int, ...]:
    result = tuple(sorted(set(ids)))
    if not result:
        raise ValueError("空人员组不能生成架次")
    return result


def fixed_airport(ids: Sequence[int], people: Sequence[Person]) -> Optional[str]:
    fixed = set()
    for person_id in ids:
        person = people[person_id]
        if person.origin in AIRPORT_SET:
            fixed.add(person.origin)
        if person.destination in AIRPORT_SET:
            fixed.add(person.destination)
    return next(iter(fixed)) if len(fixed) == 1 else None


def incompatible_fixed_airports(ids: Sequence[int], people: Sequence[Person]) -> bool:
    fixed = set()
    for person_id in ids:
        person = people[person_id]
        if person.origin in AIRPORT_SET:
            fixed.add(person.origin)
        if person.destination in AIRPORT_SET:
            fixed.add(person.destination)
    return len(fixed) > 1


def required_nodes(ids: Sequence[int], people: Sequence[Person]) -> Tuple[str, ...]:
    nodes = set()
    for person_id in ids:
        person = people[person_id]
        if person.origin not in AIRPORT_SET | {LAND}:
            nodes.add(person.origin)
        if person.destination not in AIRPORT_SET | {LAND}:
            nodes.add(person.destination)
    return tuple(sorted(nodes))


def shuttle_edges(ids: Sequence[int], people: Sequence[Person]) -> Tuple[Tuple[str, str], ...]:
    return tuple(sorted({
        people[person_id].od
        for person_id in ids
        if people[person_id].origin not in AIRPORT_SET | {LAND}
        and people[person_id].destination not in AIRPORT_SET | {LAND}
    }))


def od_counts(ids: Sequence[int], people: Sequence[Person]) -> Counter:
    return Counter(people[person_id].od for person_id in ids)


def sequence_supports(sequence: Sequence[str], edges: Sequence[Tuple[str, str]]) -> bool:
    positions: Dict[str, List[int]] = defaultdict(list)
    for position, node in enumerate(sequence):
        positions[node].append(position)
    return all(
        any(pickup < delivery for pickup in positions[origin] for delivery in positions[destination])
        for origin, destination in edges
    )


@lru_cache(maxsize=None)
def simple_sequences(
    required: Tuple[str, ...], edges: Tuple[Tuple[str, str], ...],
) -> Tuple[Tuple[str, ...], ...]:
    return tuple(
        order for order in itertools.permutations(required)
        if sequence_supports(order, edges)
    )


@lru_cache(maxsize=None)
def repeated_sequences(
    required: Tuple[str, ...], edges: Tuple[Tuple[str, str], ...],
) -> Tuple[Tuple[str, ...], ...]:
    if len(required) <= 1 or len(required) >= MAX_LANDINGS:
        return tuple()
    required_set = set(required)
    result = []
    for length in range(len(required) + 1, MAX_LANDINGS + 1):
        for sequence in itertools.product(required, repeat=length):
            if any(sequence[i] == sequence[i + 1] for i in range(length - 1)):
                continue
            if set(sequence) != required_set:
                continue
            if sequence_supports(sequence, edges):
                result.append(tuple(sequence))
    return tuple(result)


def fuel_feasible(
    airport: str,
    aircraft_type: str,
    stops: Sequence[str],
    refuels: Sequence[bool],
) -> Optional[Tuple[float, Tuple[float, ...], Tuple[int, ...], Tuple[int, ...]]]:
    ac = AC[aircraft_type]
    remain = float(ac["tank"])
    reserve = float(ac["reserve"])
    burn = float(ac["burn"])
    current = airport
    elapsed = 0
    fuel = 0.0
    arrivals = [0]
    departures = [0]
    leg_distances = []
    for stop, refuel in zip(stops, refuels):
        distance = D[current][stop]
        remain -= distance * burn
        if remain < reserve - EPS:
            return None
        elapsed += flight_min(distance, aircraft_type)
        fuel += distance * burn
        leg_distances.append(distance)
        arrivals.append(elapsed)
        if refuel and stop not in GAS_SET:
            return None
        elapsed += 20 if refuel else 10
        departures.append(elapsed)
        if refuel:
            remain = float(ac["tank"])
        current = stop
    distance = D[current][airport]
    remain -= distance * burn
    if remain < reserve - EPS:
        return None
    elapsed += flight_min(distance, aircraft_type)
    fuel += distance * burn
    leg_distances.append(distance)
    arrivals.append(elapsed)
    departures.append(elapsed)
    return fuel, tuple(leg_distances), tuple(arrivals), tuple(departures)


def refuel_variants(
    stops: Tuple[str, ...],
    forced: frozenset[int],
    airport: str,
    aircraft_type: str,
) -> List[Tuple[Tuple[bool, ...], Tuple[float, Tuple[float, ...], Tuple[int, ...], Tuple[int, ...]]]]:
    optional = [
        i for i, node in enumerate(stops)
        if node in GAS_SET and i not in forced
    ]
    candidates = []
    for bits in itertools.product((False, True), repeat=len(optional)):
        flags = [False] * len(stops)
        for i in forced:
            flags[i] = True
        for i, bit in zip(optional, bits):
            flags[i] = bit
        result = fuel_feasible(airport, aircraft_type, stops, flags)
        if result is not None:
            candidates.append((tuple(flags), result))
    if not candidates:
        return []
    minimum_refuels = min(sum(flags) for flags, _ in candidates)
    return [item for item in candidates if sum(item[0]) == minimum_refuels]


def insert_gas_paths(
    base: Tuple[str, ...], extra: int,
) -> List[Tuple[Tuple[str, ...], frozenset[int]]]:
    states = [tuple((node, False) for node in base)]
    for _ in range(extra):
        next_states = set()
        for state in states:
            for position in range(len(state) + 1):
                for gas in GAS_STATIONS:
                    before = state[position - 1][0] if position else None
                    after = state[position][0] if position < len(state) else None
                    if gas in {before, after}:
                        continue
                    next_states.add(state[:position] + ((gas, True),) + state[position:])
        states = list(next_states)
    result = []
    for state in states:
        stops = tuple(node for node, _ in state)
        forced = frozenset(i for i, (_, inserted) in enumerate(state) if inserted)
        result.append((stops, forced))
    return result


def path_variants(
    service_sequence: Tuple[str, ...], airport: str, aircraft_type: str,
) -> List[Tuple[Tuple[str, ...], Tuple[bool, ...], Tuple]]:
    direct = refuel_variants(service_sequence, frozenset(), airport, aircraft_type)
    if direct:
        return [(service_sequence, flags, result) for flags, result in direct]
    remaining = MAX_LANDINGS - len(service_sequence)
    for extra in range(1, remaining + 1):
        feasible = []
        for stops, forced in insert_gas_paths(service_sequence, extra):
            for flags, result in refuel_variants(stops, forced, airport, aircraft_type):
                feasible.append((stops, flags, result))
        if feasible:
            return feasible
    return []


def class_options(
    od: Tuple[str, str], stops: Sequence[str], airport: str,
) -> List[Tuple[int, int]]:
    origin, destination = od
    final = len(stops) + 1
    if origin in AIRPORT_SET and origin != airport:
        return []
    if destination in AIRPORT_SET and destination != airport:
        return []
    pickup_positions = [0] if origin in AIRPORT_SET | {LAND} else [
        i for i, node in enumerate(stops, start=1) if node == origin
    ]
    options = []
    for pickup in pickup_positions:
        if destination in AIRPORT_SET | {LAND}:
            options.append((pickup, final))
        else:
            later = [
                i for i, node in enumerate(stops, start=1)
                if i > pickup and node == destination
            ]
            if later:
                options.append((pickup, later[0]))
    if destination in AIRPORT_SET | {LAND} and options:
        options = [max(options, key=lambda pair: pair[0])]
    return list(dict.fromkeys(options))


def assignment_variants(
    counts: Counter,
    stops: Tuple[str, ...],
    airport: str,
    aircraft_type: str,
    leg_distances: Sequence[float],
    arrivals: Sequence[int],
    departures: Sequence[int],
    limit: int,
) -> List[Tuple[Tuple[Tuple[str, str, int, int], ...], int, float]]:
    seats = int(AC[aircraft_type]["seats"])
    entries = []
    for od, count in counts.items():
        options = class_options(od, stops, airport)
        if not options:
            return []
        entries.append((od, count, options))
    entries.sort(key=lambda item: (len(item[2]), -item[1], item[0]))
    loads = [0] * len(leg_distances)
    chosen: List[Tuple[str, str, int, int]] = []
    result = []

    def dfs(position: int, intransit: int) -> None:
        if len(result) >= limit:
            return
        if position == len(entries):
            pass_km = sum(load * distance for load, distance in zip(loads, leg_distances))
            result.append((tuple(sorted(chosen)), intransit, pass_km))
            return
        (origin, destination), count, options = entries[position]
        for pickup, delivery in sorted(
            options,
            key=lambda pair: (arrivals[pair[1]] - departures[pair[0]], pair),
        ):
            if any(loads[leg] + count > seats for leg in range(pickup, delivery)):
                continue
            for leg in range(pickup, delivery):
                loads[leg] += count
            chosen.append((origin, destination, pickup, delivery))
            dfs(
                position + 1,
                intransit + count * (arrivals[delivery] - departures[pickup]),
            )
            chosen.pop()
            for leg in range(pickup, delivery):
                loads[leg] -= count

    dfs(0, 0)
    result.sort(key=lambda item: (item[1], -item[2], item[0]))
    return result


def feasible_start_intervals(
    ids: Sequence[int],
    people: Sequence[Person],
    pairs: Dict[Tuple[str, str], Tuple[int, int]],
    arrivals: Sequence[int],
    departures: Sequence[int],
    duration: int,
) -> Tuple[Tuple[int, int], ...]:
    lower = 0
    upper = HORIZON - duration
    for person_id in ids:
        person = people[person_id]
        pickup, delivery = pairs[person.od]
        lower = max(lower, person.earliest - departures[pickup])
        upper = min(upper, person.latest - arrivals[delivery])
    if lower > upper:
        return tuple()
    intervals = []
    for day in range(PLANNING_DAYS):
        day_start = day * DAY_MINUTES
        lo = max(lower, day_start + TAKEOFF_OPEN)
        hi = min(
            upper,
            day_start + TAKEOFF_CLOSE,
            day_start + RETURN_CLOSE - duration,
        )
        if lo <= hi:
            intervals.append((int(lo), int(hi)))
    return tuple(intervals)


def route_sort_key(route: Route, intervals: Sequence[Tuple[int, int]]) -> Tuple:
    return (
        route.duration, route.intransit_min, round(route.fuel_kg, 6),
        -route.pass_km / route.avail_km,
        -sum(hi - lo + 1 for lo, hi in intervals),
        route.aircraft_type, route.airport, route.stops, route.refuels, route.pairs,
    )


class RouteFactory:
    def __init__(
        self,
        people: Sequence[Person],
        max_route_people: int,
        assignment_limit: int,
    ) -> None:
        self.people = people
        self.max_route_people = max(MAX_SEATS, max_route_people)
        self.assignment_limit = max(1, assignment_limit)
        self.cache: Dict[Tuple[int, ...], Tuple[Tuple[Route, Tuple[Tuple[int, int], ...]], ...]] = {}
        self.calls = 0
        self.cache_hits = 0

    def generate(
        self, ids: Iterable[int], max_variants: int,
    ) -> Tuple[Tuple[Route, Tuple[Tuple[int, int], ...]], ...]:
        key = canonical_people(ids)
        if key in self.cache:
            self.cache_hits += 1
            return self.cache[key][:max_variants]
        self.calls += 1
        if len(key) > self.max_route_people or incompatible_fixed_airports(key, self.people):
            self.cache[key] = tuple()
            return tuple()
        required = required_nodes(key, self.people)
        if not required or len(required) > MAX_LANDINGS:
            self.cache[key] = tuple()
            return tuple()
        fixed = fixed_airport(key, self.people)
        airports = (fixed,) if fixed else AIRPORTS
        edges = shuttle_edges(key, self.people)
        counts = od_counts(key, self.people)
        candidates: Dict[Tuple, Tuple[Route, Tuple[Tuple[int, int], ...]]] = {}

        def examine(sequences: Sequence[Tuple[str, ...]]) -> None:
            for service_sequence in sequences:
                for airport in airports:
                    for aircraft_type in AC:
                        for stops, refuels, fuel_result in path_variants(
                            service_sequence, airport, aircraft_type,
                        ):
                            fuel, leg_distances, arrivals, departures = fuel_result
                            for pairs, intransit, pass_km in assignment_variants(
                                counts, stops, airport, aircraft_type,
                                leg_distances, arrivals, departures,
                                self.assignment_limit,
                            ):
                                pair_map = {
                                    (origin, destination): (pickup, delivery)
                                    for origin, destination, pickup, delivery in pairs
                                }
                                intervals = feasible_start_intervals(
                                    key, self.people, pair_map,
                                    arrivals, departures, arrivals[-1],
                                )
                                if not intervals:
                                    continue
                                route = Route(
                                    airport=airport,
                                    aircraft_type=aircraft_type,
                                    stops=stops,
                                    refuels=refuels,
                                    pairs=pairs,
                                    duration=arrivals[-1],
                                    fuel_kg=fuel,
                                    arrivals=arrivals,
                                    departures=departures,
                                    leg_distances=leg_distances,
                                    intransit_min=intransit,
                                    pass_km=pass_km,
                                    avail_km=float(AC[aircraft_type]["seats"]) * sum(leg_distances),
                                )
                                signature = route.signature, intervals
                                candidates[signature] = (route, intervals)

        simple = simple_sequences(required, edges)
        examine(simple)
        if not candidates:
            examine(repeated_sequences(required, edges))
        ordered_all = tuple(sorted(
            candidates.values(), key=lambda item: route_sort_key(item[0], item[1])
        ))
        # 首批三个候选强制覆盖T1/T2/T3；随后覆盖不同“机场-机型”组合。
        # 否则 LAND 需求的前三条最短路线可能全是三个机场的T1，造成虚假不可行。
        diverse = []
        used = set()
        for aircraft_type in AC:
            item = next(
                (candidate for candidate in ordered_all if candidate[0].aircraft_type == aircraft_type),
                None,
            )
            if item is not None:
                signature = item[0].signature, item[1]
                if signature not in used:
                    diverse.append(item)
                    used.add(signature)
        seen_groups = {(item[0].airport, item[0].aircraft_type) for item in diverse}
        for item in ordered_all:
            group = (item[0].airport, item[0].aircraft_type)
            signature = item[0].signature, item[1]
            if group not in seen_groups and signature not in used:
                diverse.append(item)
                used.add(signature)
                seen_groups.add(group)
        diverse.extend(
            item for item in ordered_all
            if (item[0].signature, item[1]) not in used
        )
        ordered = tuple(diverse)
        self.cache[key] = ordered
        return ordered[:max_variants]


class RoutePool:
    def __init__(self, max_patterns: int) -> None:
        self.max_patterns = max(1, max_patterns)
        # active_limit 是各阶段的软上限；max_patterns 是整个程序的硬上限。
        # 软上限用于给阶段二初始生成和后续临时任务LNS预留容量。
        self.active_limit = self.max_patterns
        self.patterns: List[Pattern] = []
        self.by_signature: Dict[Tuple, int] = {}

    def __len__(self) -> int:
        return len(self.patterns)

    def set_active_limit(self, limit: int) -> int:
        self.active_limit = max(len(self.patterns), min(self.max_patterns, int(limit)))
        return self.active_limit

    @property
    def is_full(self) -> bool:
        return len(self.patterns) >= self.active_limit

    @property
    def remaining_capacity(self) -> int:
        return max(0, self.active_limit - len(self.patterns))

    def add(
        self,
        people: Iterable[int],
        route: Route,
        intervals: Tuple[Tuple[int, int], ...],
        source: str,
    ) -> Tuple[int, bool]:
        ids = canonical_people(people)
        signature = ids, route.signature, intervals
        if signature in self.by_signature:
            pattern_id = self.by_signature[signature]
            self.patterns[pattern_id].sources.add(source)
            return pattern_id, False
        if self.is_full:
            return -1, False
        pattern_id = len(self.patterns)
        pattern = Pattern(pattern_id, ids, route, intervals, {source})
        self.patterns.append(pattern)
        self.by_signature[signature] = pattern_id
        return pattern_id, True


def add_generated(
    pool: RoutePool,
    factory: RouteFactory,
    ids: Iterable[int],
    variants: int,
    source: str,
) -> int:
    added = 0
    key = canonical_people(ids)
    for route, intervals in factory.generate(key, variants):
        _, is_new = pool.add(key, route, intervals, source)
        added += int(is_new)
    return added


def build_initial_pool(
    people: Sequence[Person],
    target_ids: Sequence[int],
    factory: RouteFactory,
    pool: RoutePool,
    variants: int,
    source: str,
    add_splits: bool = True,
) -> List[Tuple[int, ...]]:
    grouped: Dict[Tuple[str, str], List[int]] = defaultdict(list)
    for person_id in target_ids:
        grouped[people[person_id].od].append(person_id)
    batches = []
    groups = sorted(
        grouped.items(),
        key=lambda item: (
            min(people[p].latest for p in item[1]),
            min(TASK_ORDER[people[p].task_type] for p in item[1]),
            item[0],
        ),
    )
    for group_number, (_, ids) in enumerate(groups, start=1):
        ids.sort(key=lambda p: (
            people[p].latest, people[p].earliest,
            TASK_ORDER[people[p].task_type], people[p].person_id,
        ))
        current: List[int] = []
        for person_id in ids:
            trial = tuple(current + [person_id])
            if len(trial) <= MAX_SEATS and factory.generate(trial, 1):
                current.append(person_id)
                continue
            if not current:
                raise RuntimeError(f"人员无法形成任何时间可行架次：{people[person_id].person_id}")
            batches.append(tuple(current))
            current = [person_id]
            if not factory.generate(current, 1):
                raise RuntimeError(f"人员无法形成任何时间可行架次：{people[person_id].person_id}")
        if current:
            batches.append(tuple(current))
        if group_number % 25 == 0 or group_number == len(groups):
            print(f"  已处理 {group_number}/{len(groups)} 个OD类别，批次={len(batches)}")

    for batch in batches:
        candidates = factory.generate(batch, variants)
        if not candidates:
            raise RuntimeError(f"初始批次没有候选路线：{[people[p].person_id for p in batch[:3]]}")
        for route, intervals in candidates:
            pool.add(batch, route, intervals, source)
        if add_splits and len(batch) >= 4:
            midpoint = len(batch) // 2
            for part in (batch[:midpoint], batch[midpoint:]):
                add_generated(pool, factory, part, variants, f"{source}_split")
                if len(part) >= 8:
                    quarter = len(part) // 2
                    add_generated(pool, factory, part[:quarter], variants, f"{source}_split2")
                    add_generated(pool, factory, part[quarter:], variants, f"{source}_split2")
    return batches


def scaled(value: float) -> int:
    return int(round(value * SCALE))


def metrics_from_selections(
    selections: Sequence[Selection],
    pool: RoutePool,
    people: Sequence[Person],
) -> Metrics:
    temporary = 0
    aircraft_time = 0
    intransit = 0
    fuel = 0.0
    pass_km = 0.0
    avail_km = 0.0
    for selection in selections:
        pattern = pool.patterns[selection.pattern_id]
        route = pattern.route
        temporary += sum(people[p].task_type == "temporary" for p in pattern.people)
        aircraft_time += route.duration
        intransit += route.intransit_min
        fuel += route.fuel_kg
        pass_km += route.pass_km
        avail_km += route.avail_km
    return Metrics(
        temporary_served=temporary,
        aircraft_time=aircraft_time,
        person_intransit=intransit,
        sorties=len(selections),
        fuel_kg=fuel,
        pass_km=pass_km,
        avail_km=avail_km,
    )


def stage1_better(a: Metrics, b: Optional[Metrics]) -> bool:
    if b is None:
        return True
    return (
        a.aircraft_time, a.person_intransit, a.sorties,
        round(a.fuel_kg, 6), -a.seat_utilization,
    ) < (
        b.aircraft_time, b.person_intransit, b.sorties,
        round(b.fuel_kg, 6), -b.seat_utilization,
    )


def stage2_better(a: Metrics, b: Optional[Metrics]) -> bool:
    if b is None:
        return True
    return (
        -a.temporary_served, a.aircraft_time, a.person_intransit,
        a.sorties, round(a.fuel_kg, 6), -a.seat_utilization,
    ) < (
        -b.temporary_served, b.aircraft_time, b.person_intransit,
        b.sorties, round(b.fuel_kg, 6), -b.seat_utilization,
    )


def solve_master(
    pool: RoutePool,
    people: Sequence[Person],
    mandatory_ids: Sequence[int],
    temporary_ids: Sequence[int],
    stage: int,
    baseline_cap: Optional[int],
    incumbent: Optional[Sequence[Selection]],
    time_limit: float,
    workers: int,
    seed: int,
) -> MasterResult:
    if stage not in (1, 2):
        raise ValueError("stage必须为1或2")
    mandatory_set = set(mandatory_ids)
    temporary_set = set(temporary_ids)
    eligible = [
        pattern for pattern in pool.patterns
        if stage == 2 or all(person_id in mandatory_set for person_id in pattern.people)
    ]
    if not eligible:
        raise RuntimeError("候选路线池为空")
    eligible_ids = {pattern.pattern_id for pattern in eligible}
    model = cp_model.CpModel()
    selected_var = {
        pattern.pattern_id: model.NewBoolVar(f"z_{pattern.pattern_id}")
        for pattern in eligible
    }
    patterns_by_person: Dict[int, List[int]] = defaultdict(list)
    for pattern in eligible:
        for person_id in pattern.people:
            patterns_by_person[person_id].append(pattern.pattern_id)
    for person_id in mandatory_ids:
        terms = [selected_var[p] for p in patterns_by_person.get(person_id, [])]
        if not terms:
            raise RuntimeError(f"路线池未覆盖必需人员：{people[person_id].person_id}")
        model.Add(sum(terms) == 1)
    if stage == 2:
        for person_id in temporary_ids:
            terms = [selected_var[p] for p in patterns_by_person.get(person_id, [])]
            if terms:
                model.Add(sum(terms) <= 1)

    assignment_vars = {}
    intervals_by_aircraft: Dict[str, List] = defaultdict(list)
    for pattern in eligible:
        route = pattern.route
        choices = []
        domain = cp_model.Domain.FromIntervals([
            [lo, hi] for lo, hi in pattern.start_intervals
        ])
        for aircraft_id in FLEET_BY_GROUP[(route.airport, route.aircraft_type)]:
            present = model.NewBoolVar(f"use_{pattern.pattern_id}_{aircraft_id}")
            start_var = model.NewIntVarFromDomain(
                domain, f"start_{pattern.pattern_id}_{aircraft_id}"
            )
            end_var = model.NewIntVar(
                0, HORIZON + TURNAROUND,
                f"block_end_{pattern.pattern_id}_{aircraft_id}",
            )
            interval = model.NewOptionalIntervalVar(
                start_var, route.duration + TURNAROUND, end_var, present,
                f"interval_{pattern.pattern_id}_{aircraft_id}",
            )
            assignment_vars[(pattern.pattern_id, aircraft_id)] = (present, start_var)
            intervals_by_aircraft[aircraft_id].append(interval)
            choices.append(present)
        model.Add(sum(choices) == selected_var[pattern.pattern_id])
    for aircraft_id, intervals in intervals_by_aircraft.items():
        model.AddNoOverlap(intervals)

    time_expr = sum(
        pattern.route.duration * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    transit_expr = sum(
        pattern.route.intransit_min * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    sorties_expr = sum(selected_var.values())
    fuel_expr = sum(
        scaled(pattern.route.fuel_kg) * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    pass_expr = sum(
        scaled(pattern.route.pass_km) * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    avail_expr = sum(
        scaled(pattern.route.avail_km) * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    temporary_expr = sum(
        sum(person_id in temporary_set for person_id in pattern.people)
        * selected_var[pattern.pattern_id]
        for pattern in eligible
    )
    if baseline_cap is not None:
        model.Add(time_expr <= baseline_cap)

    incumbent_map = {
        selection.pattern_id: selection for selection in (incumbent or [])
        if selection.pattern_id in eligible_ids
    }
    for pattern in eligible:
        model.AddHint(selected_var[pattern.pattern_id], int(pattern.pattern_id in incumbent_map))
        incumbent_selection = incumbent_map.get(pattern.pattern_id)
        for aircraft_id in FLEET_BY_GROUP[(pattern.route.airport, pattern.route.aircraft_type)]:
            present, start_var = assignment_vars[(pattern.pattern_id, aircraft_id)]
            is_match = bool(
                incumbent_selection and incumbent_selection.aircraft_id == aircraft_id
            )
            model.AddHint(present, int(is_match))
            if is_match:
                model.AddHint(start_var, incumbent_selection.start)

    deadline = time.monotonic() + max(0.1, time_limit)
    stages = []
    best_selections = list(incumbent or [])
    all_optimal = True

    def read_solution(solver: cp_model.CpSolver) -> List[Selection]:
        result = []
        for pattern in eligible:
            if not solver.Value(selected_var[pattern.pattern_id]):
                continue
            assigned = False
            for aircraft_id in FLEET_BY_GROUP[(pattern.route.airport, pattern.route.aircraft_type)]:
                present, start_var = assignment_vars[(pattern.pattern_id, aircraft_id)]
                if solver.Value(present):
                    result.append(Selection(
                        pattern.pattern_id, aircraft_id, solver.Value(start_var)
                    ))
                    assigned = True
                    break
            if not assigned:
                raise RuntimeError(f"被选路线没有具体飞机：{pattern.pattern_id}")
        return result

    def run(expr, maximize: bool, name: str):
        remaining = deadline - time.monotonic()
        if remaining <= 0.05:
            stages.append({"stage": name, "status": "SKIPPED_TIME_LIMIT"})
            return None, None
        model.Maximize(expr) if maximize else model.Minimize(expr)
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = max(0.05, remaining)
        solver.parameters.num_search_workers = max(1, workers)
        solver.parameters.random_seed = seed
        solver.parameters.log_search_progress = False
        t0 = time.monotonic()
        status = solver.Solve(model)
        info = {
            "stage": name,
            "status": solver.StatusName(status),
            "wall_time_s": round(time.monotonic() - t0, 3),
        }
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            info["objective"] = int(round(solver.ObjectiveValue()))
        stages.append(info)
        return solver, status

    objectives = []
    if stage == 2:
        objectives.append(("temporary_served", temporary_expr, True))
    objectives.extend([
        ("total_aircraft_time", time_expr, False),
        ("total_person_intransit", transit_expr, False),
        ("total_sorties", sorties_expr, False),
        ("total_fuel_scaled", fuel_expr, False),
    ])
    completed_linear = True
    for name, expr, maximize in objectives:
        solver, status = run(expr, maximize, name)
        if solver is None or status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            all_optimal = False
            completed_linear = False
            break
        best_selections = read_solution(solver)
        if status != cp_model.OPTIMAL:
            all_optimal = False
            completed_linear = False
            break
        model.Add(expr == int(round(solver.ObjectiveValue())))

    if completed_linear:
        current_metrics = metrics_from_selections(best_selections, pool, people)
        current_pass = scaled(current_metrics.pass_km)
        current_avail = scaled(current_metrics.avail_km)
        ratio_optimal = False
        if current_avail > 0:
            for iteration in range(12):
                divisor = math.gcd(abs(current_pass), abs(current_avail)) or 1
                q_num = current_pass // divisor
                q_den = current_avail // divisor
                solver, status = run(
                    q_den * pass_expr - q_num * avail_expr,
                    True,
                    f"seat_utilization_{iteration + 1}",
                )
                if solver is None or status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
                    all_optimal = False
                    break
                best_selections = read_solution(solver)
                metrics = metrics_from_selections(best_selections, pool, people)
                next_pass = scaled(metrics.pass_km)
                next_avail = scaled(metrics.avail_km)
                if status != cp_model.OPTIMAL:
                    all_optimal = False
                    break
                residual = q_den * next_pass - q_num * next_avail
                current_pass, current_avail = next_pass, next_avail
                if residual == 0:
                    ratio_optimal = True
                    break
        if not ratio_optimal:
            all_optimal = False

    if not best_selections:
        raise RuntimeError(
            "限定时间内没有得到可行排班。建议提高 --master-time，或增加机型候选。"
            f" 求解阶段={stages}"
        )
    metrics = metrics_from_selections(best_selections, pool, people)
    return MasterResult(best_selections, metrics, stages, all_optimal)


def route_nodes(pattern: Pattern) -> set[str]:
    return set(pattern.route.stops) | {pattern.route.airport}


def batch_nodes(batch: Sequence[int], people: Sequence[Person]) -> set[str]:
    nodes = set()
    for person_id in batch:
        person = people[person_id]
        for node in person.od:
            if node not in AIRPORT_SET | {LAND}:
                nodes.add(node)
    return nodes


def batch_pair_score(
    a: Sequence[int], b: Sequence[int], people: Sequence[Person], max_people: int,
) -> Optional[Tuple]:
    if len(a) + len(b) > max_people:
        return None
    union = tuple(a) + tuple(b)
    if incompatible_fixed_airports(union, people):
        return None
    nodes_a = batch_nodes(a, people)
    nodes_b = batch_nodes(b, people)
    if len(nodes_a | nodes_b) > MAX_LANDINGS:
        return None
    distance = min(D[x][y] for x in nodes_a for y in nodes_b)
    earliest_gap = abs(
        min(people[p].earliest for p in a) - min(people[p].earliest for p in b)
    )
    deadline_gap = abs(
        min(people[p].latest for p in a) - min(people[p].latest for p in b)
    )
    shared = len(nodes_a & nodes_b)
    return (deadline_gap, earliest_gap, distance, -shared, -(len(a) + len(b)))


def bootstrap_merge_batches(
    batches: Sequence[Tuple[int, ...]],
    people: Sequence[Person],
    pool: RoutePool,
    factory: RouteFactory,
    variants: int,
    neighbors: int,
    triple_attempts: int,
    rng: random.Random,
    deadline: float,
) -> int:
    added = 0
    feasible_neighbors: Dict[int, List[int]] = defaultdict(list)
    considered = set()
    for i, batch in enumerate(batches):
        if time.monotonic() >= deadline or pool.is_full:
            break
        ranked = []
        for j, other in enumerate(batches):
            if i == j:
                continue
            score = batch_pair_score(batch, other, people, factory.max_route_people)
            if score is not None:
                ranked.append((score, j))
        ranked.sort(key=lambda item: item[0])
        for _, j in ranked[:max(1, neighbors)]:
            pair = tuple(sorted((i, j)))
            if pair in considered:
                continue
            considered.add(pair)
            union = tuple(sorted(batches[i] + batches[j]))
            count = add_generated(pool, factory, union, variants, "bootstrap_merge2")
            if count:
                added += count
                feasible_neighbors[i].append(j)
                feasible_neighbors[j].append(i)
        if (i + 1) % 50 == 0:
            print(f"  bootstrap {i + 1}/{len(batches)}，新增路线={added}")

    candidate_centers = [i for i, values in feasible_neighbors.items() if len(values) >= 2]
    for _ in range(max(0, triple_attempts)):
        if (
            not candidate_centers
            or time.monotonic() >= deadline
            or pool.is_full
        ):
            break
        center = rng.choice(candidate_centers)
        first, second = rng.sample(feasible_neighbors[center], 2)
        union = tuple(sorted(
            set(batches[center]) | set(batches[first]) | set(batches[second])
        ))
        if len(union) > factory.max_route_people:
            continue
        added += add_generated(pool, factory, union, variants, "bootstrap_merge3")
    return added


def pattern_proximity(a: Pattern, b: Pattern) -> Tuple:
    nodes_a = route_nodes(a)
    nodes_b = route_nodes(b)
    distance = min(D[x][y] for x in nodes_a for y in nodes_b)
    common_days = 0
    for lo_a, hi_a in a.start_intervals:
        for lo_b, hi_b in b.start_intervals:
            common_days += int(max(lo_a, lo_b) <= min(hi_a, hi_b))
    return (
        -common_days, distance,
        abs(a.route.duration - b.route.duration),
        a.pattern_id, b.pattern_id,
    )


def expand_selected_pool(
    selections: Sequence[Selection],
    pool: RoutePool,
    factory: RouteFactory,
    variants: int,
    neighbors: int,
    attempts: int,
    rng: random.Random,
    deadline: float,
    source: str,
) -> int:
    active = [pool.patterns[selection.pattern_id] for selection in selections]
    added = 0
    considered = set()
    for i, pattern in enumerate(active):
        if time.monotonic() >= deadline or pool.is_full:
            break
        ranked = sorted(
            (
                (pattern_proximity(pattern, other), j)
                for j, other in enumerate(active)
                if i != j
            ),
            key=lambda item: item[0],
        )
        for _, j in ranked[:max(1, neighbors)]:
            pair = tuple(sorted((i, j)))
            if pair in considered:
                continue
            considered.add(pair)
            union = tuple(sorted(set(pattern.people) | set(active[j].people)))
            if len(union) > factory.max_route_people:
                continue
            added += add_generated(pool, factory, union, variants, f"{source}_merge")

    if len(active) < 2:
        return added
    for _ in range(max(0, attempts)):
        if time.monotonic() >= deadline or pool.is_full:
            break
        a, b = rng.sample(active, 2)
        union = tuple(sorted(set(a.people) | set(b.people)))
        if len(union) < 3 or len(union) > 2 * factory.max_route_people:
            continue
        ordered = sorted(
            union,
            key=lambda p: (
                factory.people[p].latest,
                factory.people[p].earliest,
                factory.people[p].origin,
                factory.people[p].destination,
            ),
        )
        strategies = []
        midpoint = len(ordered) // 2
        strategies.append((tuple(ordered[:midpoint]), tuple(ordered[midpoint:])))
        strategies.append((tuple(ordered[::2]), tuple(ordered[1::2])))
        by_od = sorted(
            union,
            key=lambda p: (
                factory.people[p].origin,
                factory.people[p].destination,
                factory.people[p].latest,
            ),
        )
        strategies.append((tuple(by_od[:midpoint]), tuple(by_od[midpoint:])))
        first, second = rng.choice(strategies)
        if not first or not second:
            continue
        before = len(pool)
        first_added = add_generated(pool, factory, first, variants, f"{source}_repartition")
        second_added = add_generated(pool, factory, second, variants, f"{source}_repartition")
        if first_added and second_added:
            added += len(pool) - before
    return added


def person_route_distance(person: Person, pattern: Pattern) -> float:
    nodes = route_nodes(pattern)
    endpoints = [
        node for node in (person.origin, person.destination)
        if node not in AIRPORT_SET | {LAND}
    ]
    fixed_penalty = 0.0
    if person.origin in AIRPORT_SET and person.origin != pattern.route.airport:
        return float("inf")
    if person.destination in AIRPORT_SET and person.destination != pattern.route.airport:
        return float("inf")
    if not endpoints:
        fixed_penalty = 1000.0
    return fixed_penalty + sum(min(D[node][other] for other in nodes) for node in endpoints)


def expand_temporary_pool(
    selections: Sequence[Selection],
    temporary_ids: Sequence[int],
    pool: RoutePool,
    factory: RouteFactory,
    variants: int,
    candidates_per_route: int,
    deadline: float,
    source: str,
) -> int:
    served = {
        person_id
        for selection in selections
        for person_id in pool.patterns[selection.pattern_id].people
        if factory.people[person_id].task_type == "temporary"
    }
    unserved = [person_id for person_id in temporary_ids if person_id not in served]
    added = 0
    for selection in selections:
        if time.monotonic() >= deadline or pool.is_full:
            break
        pattern = pool.patterns[selection.pattern_id]
        room = factory.max_route_people - len(pattern.people)
        if room <= 0:
            continue
        ranked = sorted(
            unserved,
            key=lambda p: (
                person_route_distance(factory.people[p], pattern),
                factory.people[p].latest,
                factory.people[p].earliest,
            ),
        )
        feasible_singles = []
        for person_id in ranked[:max(1, candidates_per_route * 3)]:
            if math.isinf(person_route_distance(factory.people[person_id], pattern)):
                continue
            union = tuple(sorted(pattern.people + (person_id,)))
            count = add_generated(pool, factory, union, variants, f"{source}_insert1")
            if count:
                feasible_singles.append(person_id)
                added += count
            if len(feasible_singles) >= candidates_per_route:
                break
        for first, second in itertools.combinations(feasible_singles[:6], 2):
            if time.monotonic() >= deadline or room < 2:
                break
            union = tuple(sorted(pattern.people + (first, second)))
            added += add_generated(pool, factory, union, variants, f"{source}_insert2")
    return added


def selection_signature(selections: Sequence[Selection]) -> Tuple:
    return tuple(sorted(
        (selection.pattern_id, selection.aircraft_id, selection.start)
        for selection in selections
    ))


def print_metrics(label: str, metrics: Metrics, pool_size: int) -> None:
    print(label)
    print(f"  临时人员完成数:   {metrics.temporary_served}")
    print(f"  总飞机使用时间:   {metrics.aircraft_time} min ({metrics.aircraft_time / 60:.2f} h)")
    print(f"  人员总在途时间:   {metrics.person_intransit} min")
    print(f"  总架次数:         {metrics.sorties}")
    print(f"  总燃油:           {metrics.fuel_kg:.2f} kg")
    print(f"  座位利用率:       {metrics.seat_utilization:.6f}")
    print(f"  路线池规模:       {pool_size}")


def actual_node(route: Route, stop_order: int) -> str:
    if stop_order == 0 or stop_order == len(route.stops) + 1:
        return route.airport
    return route.stops[stop_order - 1]


def validate_solution(
    selections: Sequence[Selection],
    pool: RoutePool,
    people: Sequence[Person],
    baseline_cap: Optional[int],
) -> Metrics:
    coverage = Counter()
    aircraft_calendar: Dict[str, List[Tuple[int, int, int]]] = defaultdict(list)
    for selection in selections:
        if selection.pattern_id < 0 or selection.pattern_id >= len(pool.patterns):
            raise ValueError("选择结果含未知路线编号")
        pattern = pool.patterns[selection.pattern_id]
        route = pattern.route
        airport, aircraft_type = AIRCRAFT_INFO.get(selection.aircraft_id, (None, None))
        if (airport, aircraft_type) != (route.airport, route.aircraft_type):
            raise ValueError(f"飞机与路线不兼容：{selection.aircraft_id}")
        if not any(lo <= selection.start <= hi for lo, hi in pattern.start_intervals):
            raise ValueError(f"起飞时刻不在候选可行域：pattern={pattern.pattern_id}")
        day = selection.start // DAY_MINUTES
        if day < 0 or day >= PLANNING_DAYS:
            raise ValueError("架次日期超出规划期")
        minute_of_day = selection.start - day * DAY_MINUTES
        if not TAKEOFF_OPEN <= minute_of_day <= TAKEOFF_CLOSE:
            raise ValueError("架次起飞不在06:00—18:00")
        if selection.start + route.duration > day * DAY_MINUTES + RETURN_CLOSE:
            raise ValueError("架次未在20:00前返场")
        if len(route.stops) > MAX_LANDINGS:
            raise ValueError("海上着陆次数超过5")
        if len(route.stops) != len(route.refuels):
            raise ValueError("停靠与加油标志长度不一致")

        fuel_result = fuel_feasible(
            route.airport, route.aircraft_type, route.stops, route.refuels,
        )
        if fuel_result is None:
            raise ValueError("最终路线燃油不可行")
        fuel, leg_distances, arrivals, departures = fuel_result
        if (
            arrivals != route.arrivals
            or departures != route.departures
            or leg_distances != route.leg_distances
            or abs(fuel - route.fuel_kg) > 1e-6
        ):
            raise ValueError("最终路线缓存指标与重新计算不一致")

        pair_map = route.pair_map
        loads = [0] * len(route.leg_distances)
        recomputed_intransit = 0
        recomputed_pass_km = 0.0
        for person_id in pattern.people:
            person = people[person_id]
            coverage[person_id] += 1
            if person.od not in pair_map:
                raise ValueError(f"路线缺少人员OD：{person.person_id}")
            pickup, delivery = pair_map[person.od]
            if not 0 <= pickup < delivery <= len(route.stops) + 1:
                raise ValueError(f"上下机序号非法：{person.person_id}")
            pickup_node = actual_node(route, pickup)
            delivery_node = actual_node(route, delivery)
            if person.origin == LAND:
                if pickup != 0:
                    raise ValueError(f"LAND出发人员未在机场上机：{person.person_id}")
            elif pickup_node != person.origin:
                raise ValueError(f"人员上机地点错误：{person.person_id}")
            if person.destination == LAND:
                if delivery != len(route.stops) + 1:
                    raise ValueError(f"LAND到达人员未在机场下机：{person.person_id}")
            elif delivery_node != person.destination:
                raise ValueError(f"人员下机地点错误：{person.person_id}")
            if person.destination not in AIRPORT_SET | {LAND}:
                first_destination = next(
                    (
                        position for position in range(pickup + 1, len(route.stops) + 1)
                        if actual_node(route, position) == person.destination
                    ),
                    None,
                )
                if delivery != first_destination:
                    raise ValueError(f"未在首次到达目的设施时下机：{person.person_id}")
            pickup_time = selection.start + route.departures[pickup]
            delivery_time = selection.start + route.arrivals[delivery]
            if pickup_time < person.earliest:
                raise ValueError(f"早于最早可离开时刻：{person.person_id}")
            if delivery_time > person.latest:
                raise ValueError(f"晚于最晚到达时刻：{person.person_id}")
            recomputed_intransit += delivery_time - pickup_time
            for leg in range(pickup, delivery):
                loads[leg] += 1
        seats = int(AC[route.aircraft_type]["seats"])
        if any(load > seats for load in loads):
            raise ValueError(f"动态载客量超过座位数：pattern={pattern.pattern_id}")
        recomputed_pass_km = sum(
            load * distance for load, distance in zip(loads, route.leg_distances)
        )
        if recomputed_intransit != route.intransit_min:
            raise ValueError("人员在途时间缓存错误")
        if abs(recomputed_pass_km - route.pass_km) > 1e-6:
            raise ValueError("客公里缓存错误")
        aircraft_calendar[selection.aircraft_id].append((
            selection.start,
            selection.start + route.duration,
            pattern.pattern_id,
        ))

    for person in people:
        count = coverage[person.index]
        if person.mandatory and count != 1:
            raise ValueError(f"必需人员未恰好服务一次：{person.person_id}, count={count}")
        if not person.mandatory and count not in (0, 1):
            raise ValueError(f"临时人员重复服务：{person.person_id}")
    for aircraft_id, flights in aircraft_calendar.items():
        flights.sort()
        for previous, current in zip(flights, flights[1:]):
            if current[0] < previous[1] + TURNAROUND:
                raise ValueError(
                    f"同一飞机架次重叠或周转不足：{aircraft_id}, "
                    f"pattern {previous[2]} -> {current[2]}"
                )
    metrics = metrics_from_selections(selections, pool, people)
    if baseline_cap is not None and metrics.aircraft_time > baseline_cap:
        raise ValueError(
            f"第二阶段总飞机时间超过基准：{metrics.aircraft_time}>{baseline_cap}"
        )
    return metrics


def write_outputs(
    selections: Sequence[Selection],
    pool: RoutePool,
    people: Sequence[Person],
    routes_path: Path,
    assignments_path: Path,
) -> None:
    routes_path.parent.mkdir(parents=True, exist_ok=True)
    assignments_path.parent.mkdir(parents=True, exist_ok=True)
    ordered = sorted(selections, key=lambda s: (s.aircraft_id, s.start, s.pattern_id))
    flight_counter = Counter()
    flight_number: Dict[Selection, int] = {}
    for selection in ordered:
        flight_counter[selection.aircraft_id] += 1
        flight_number[selection] = flight_counter[selection.aircraft_id]

    with routes_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "aircraft_id", "flight_no", "stop_order", "facility_id",
            "arrival_time", "departure_time", "refuel",
        ])
        for selection in ordered:
            route = pool.patterns[selection.pattern_id].route
            number = flight_number[selection]
            writer.writerow([
                selection.aircraft_id, number, 0, route.airport,
                "", format_minute(selection.start), 0,
            ])
            for stop_order, (stop, refuel) in enumerate(
                zip(route.stops, route.refuels), start=1,
            ):
                writer.writerow([
                    selection.aircraft_id, number, stop_order, stop,
                    format_minute(selection.start + route.arrivals[stop_order]),
                    format_minute(selection.start + route.departures[stop_order]),
                    int(refuel),
                ])
            writer.writerow([
                selection.aircraft_id, number, len(route.stops) + 1, route.airport,
                format_minute(selection.start + route.duration), "", 0,
            ])

    person_assignment = {}
    for selection in ordered:
        pattern = pool.patterns[selection.pattern_id]
        pair_map = pattern.route.pair_map
        number = flight_number[selection]
        for person_id in pattern.people:
            pickup, delivery = pair_map[people[person_id].od]
            person_assignment[person_id] = (
                selection.aircraft_id, number, pickup, delivery,
            )
    with assignments_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "person_id", "aircraft_id", "flight_no",
            "pickup_stop_order", "delivery_stop_order",
        ])
        for person in people:
            assignment = person_assignment.get(person.index)
            if assignment is None:
                if person.mandatory:
                    raise RuntimeError(f"写出时发现必需人员未安排：{person.person_id}")
                writer.writerow([person.person_id, "", "", "", ""])
            else:
                writer.writerow([person.person_id, *assignment])


def configure_chinese_plot() -> None:
    """配置论文图使用的中文字体。"""
    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["savefig.dpi"] = 300


def report_path(path: Path) -> Path:
    """相对路径统一相对于项目根目录。"""
    path = Path(path)
    return path if path.is_absolute() else BASE_DIR / path


def demand_kind(person: Person) -> str:
    """划分出海、海返和设施穿梭需求。"""
    land_nodes = AIRPORT_SET | {LAND}
    if person.origin in land_nodes and person.destination not in land_nodes:
        return "出海"
    if person.origin not in land_nodes and person.destination in land_nodes:
        return "海返"
    if person.origin not in land_nodes and person.destination not in land_nodes:
        return "设施穿梭"
    return "陆地间"


def window_service_days(person: Person) -> int:
    """统计时间窗与每日运行区间相交的日期数。"""
    count = 0
    for day in range(PLANNING_DAYS):
        lower = max(person.earliest, day * DAY_MINUTES + TAKEOFF_OPEN)
        upper = min(person.latest, day * DAY_MINUTES + RETURN_CLOSE)
        count += int(lower <= upper)
    return count


def read_result_frames(
    routes_path: Path,
    assignments_path: Path,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """读取现存排班结果，不调用优化器。"""
    routes = pd.read_csv(routes_path, encoding="utf-8-sig")
    assignments = pd.read_csv(assignments_path, encoding="utf-8-sig")
    route_columns = {
        "aircraft_id", "flight_no", "stop_order", "facility_id",
        "arrival_time", "departure_time", "refuel",
    }
    assignment_columns = {
        "person_id", "aircraft_id", "flight_no",
        "pickup_stop_order", "delivery_stop_order",
    }
    if not route_columns.issubset(routes.columns):
        raise ValueError("q3-routes.csv字段不完整")
    if not assignment_columns.issubset(assignments.columns):
        raise ValueError("q3-assignments.csv字段不完整")
    routes["flight_no"] = routes["flight_no"].astype(int)
    routes["stop_order"] = routes["stop_order"].astype(int)
    routes["refuel"] = routes["refuel"].fillna(0).astype(int)
    routes["arrival_dt"] = pd.to_datetime(routes["arrival_time"], errors="coerce")
    routes["departure_dt"] = pd.to_datetime(routes["departure_time"], errors="coerce")
    served = assignments.dropna(subset=[
        "aircraft_id", "flight_no", "pickup_stop_order", "delivery_stop_order",
    ]).copy()
    for column in ("flight_no", "pickup_stop_order", "delivery_stop_order"):
        served[column] = served[column].astype(int)
    return routes, served


def rebuild_flight_analysis(
    routes: pd.DataFrame,
    served_assignments: pd.DataFrame,
    people: Sequence[Person],
) -> pd.DataFrame:
    """由结果CSV重建逐架次时间、载客、燃油和利用率。"""
    people_by_id = {person.person_id: person for person in people}
    rows = []
    for (aircraft_id, flight_no), group in routes.groupby(
        ["aircraft_id", "flight_no"], sort=True,
    ):
        stops = group.sort_values("stop_order").reset_index(drop=True)
        orders = stops["stop_order"].tolist()
        if orders != list(range(len(stops))):
            raise ValueError(f"架次停靠序号不连续：{aircraft_id}-{flight_no}")
        nodes = stops["facility_id"].astype(str).tolist()
        if nodes[0] != nodes[-1]:
            raise ValueError(f"架次未返回原机场：{aircraft_id}-{flight_no}")
        start_dt = stops.iloc[0]["departure_dt"]
        return_dt = stops.iloc[-1]["arrival_dt"]
        if pd.isna(start_dt) or pd.isna(return_dt):
            raise ValueError(f"架次首尾时刻缺失：{aircraft_id}-{flight_no}")
        aircraft_info = AIRCRAFT_INFO.get(str(aircraft_id))
        if aircraft_info is None:
            raise ValueError(f"未知具体飞机：{aircraft_id}")
        airport, aircraft_type = aircraft_info
        if airport != nodes[0]:
            raise ValueError(f"具体飞机与起飞机场不一致：{aircraft_id}-{flight_no}")
        leg_distances = [D[nodes[index]][nodes[index + 1]] for index in range(len(nodes) - 1)]
        assignment_group = served_assignments[
            (served_assignments["aircraft_id"] == aircraft_id)
            & (served_assignments["flight_no"] == flight_no)
        ]
        loads = np.zeros(len(leg_distances), dtype=int)
        person_intransit = 0
        route_people = []
        arrivals = stops.set_index("stop_order")["arrival_dt"].to_dict()
        departures = stops.set_index("stop_order")["departure_dt"].to_dict()
        for assignment in assignment_group.itertuples(index=False):
            person = people_by_id.get(str(assignment.person_id))
            if person is None:
                raise ValueError(f"结果含未知人员：{assignment.person_id}")
            pickup = int(assignment.pickup_stop_order)
            delivery = int(assignment.delivery_stop_order)
            if not 0 <= pickup < delivery < len(nodes):
                raise ValueError(f"上下机序号非法：{assignment.person_id}")
            loads[pickup:delivery] += 1
            pickup_time = departures.get(pickup)
            delivery_time = arrivals.get(delivery)
            if pd.isna(pickup_time) or pd.isna(delivery_time):
                raise ValueError(f"人员上下机时刻缺失：{assignment.person_id}")
            person_intransit += int((delivery_time - pickup_time).total_seconds() // 60)
            route_people.append(person)
        duration = int((return_dt - start_dt).total_seconds() // 60)
        pass_km = float(np.dot(loads, np.asarray(leg_distances)))
        avail_km = float(AC[aircraft_type]["seats"] * sum(leg_distances))
        day_index = int((start_dt.to_pydatetime() - HORIZON_START).total_seconds() // 60) // DAY_MINUTES
        rows.append({
            "aircraft_id": str(aircraft_id),
            "flight_no": int(flight_no),
            "date": start_dt.strftime("%Y-%m-%d"),
            "day_index": day_index,
            "airport": airport,
            "aircraft_type": aircraft_type,
            "start_minute": int((start_dt.to_pydatetime() - HORIZON_START).total_seconds() // 60),
            "return_minute": int((return_dt.to_pydatetime() - HORIZON_START).total_seconds() // 60),
            "turnaround_end_minute": int((return_dt.to_pydatetime() - HORIZON_START).total_seconds() // 60) + TURNAROUND,
            "departure_time": start_dt,
            "return_time": return_dt,
            "aircraft_time_min": duration,
            "person_intransit_min": person_intransit,
            "people_count": len(route_people),
            "mandatory_count": sum(person.mandatory for person in route_people),
            "temporary_count": sum(not person.mandatory for person in route_people),
            "fuel_kg": float(AC[aircraft_type]["burn"] * sum(leg_distances)),
            "pass_km": pass_km,
            "avail_km": avail_km,
            "seat_utilization": pass_km / avail_km if avail_km > EPS else 0.0,
            "sea_landings": len(nodes) - 2,
            "refuel_count": int(stops.iloc[1:-1]["refuel"].sum()),
            "has_shuttle": any(demand_kind(person) == "设施穿梭" for person in route_people),
        })
    return pd.DataFrame(rows)


def verify_rebuilt_metrics(
    flights: pd.DataFrame,
    served_assignments: pd.DataFrame,
    people: Sequence[Person],
    summary: dict,
) -> None:
    """核对CSV复算指标与上次summary记录一致。"""
    people_by_id = {person.person_id: person for person in people}
    temporary_served = sum(
        not people_by_id[str(person_id)].mandatory
        for person_id in served_assignments["person_id"]
    )
    pass_km = float(flights["pass_km"].sum())
    avail_km = float(flights["avail_km"].sum())
    actual = {
        "temporary_served": temporary_served,
        "total_aircraft_time_min": int(flights["aircraft_time_min"].sum()),
        "total_person_intransit_min": int(flights["person_intransit_min"].sum()),
        "sorties": len(flights),
        "total_fuel_kg": float(flights["fuel_kg"].sum()),
        "seat_utilization": pass_km / avail_km if avail_km > EPS else 0.0,
    }
    expected = summary.get("final_metrics", {})
    for key, value in actual.items():
        if key not in expected:
            continue
        tolerance = 1.0e-5 if isinstance(value, float) else 0
        if abs(float(value) - float(expected[key])) > tolerance:
            raise ValueError(
                f"结果文件与summary不一致：{key}, CSV复算={value}, JSON={expected[key]}"
            )


def build_parameter_table(args: argparse.Namespace) -> pd.DataFrame:
    """记录本次最优结果对应的搜索参数。"""
    rows = [
        ("time-limit", "总搜索时间/s", args.time_limit),
        ("stage1-fraction", "阶段一时间比例", args.stage1_fraction),
        ("master-time", "单次Master时间/s", args.master_time),
        ("epochs", "阶段一LNS轮数", args.epochs),
        ("temp-epochs", "阶段二LNS轮数", args.temp_epochs),
        ("ops-per-epoch", "每轮重分包次数", args.ops_per_epoch),
        ("neighbors", "路线邻居数", args.neighbors),
        ("bootstrap-neighbors", "初始合并邻居数", args.bootstrap_neighbors),
        ("bootstrap-triples", "三路线合并尝试数", args.bootstrap_triples),
        ("temp-candidates-per-route", "每条路线临时人员候选数", args.temp_candidates_per_route),
        ("route-variants", "每个人员组合保留路线数", args.route_variants),
        ("max-patterns", "路线池上限", args.max_patterns),
        ("stage1-pool-fraction", "阶段一路线池比例", args.stage1_pool_fraction),
        ("stage2-bootstrap-pool-fraction", "阶段二初始路线池比例", args.stage2_bootstrap_pool_fraction),
        ("max-route-people", "单次Oracle最大候选人数", args.max_route_people),
        ("assignment-limit", "上下机方案枚举上限", args.assignment_limit),
        ("workers", "并行工作线程数", args.workers),
        ("seed", "随机种子（未显式输入时为默认值）", args.seed),
    ]
    return pd.DataFrame(rows, columns=["参数", "含义", "本次取值"])


def build_demand_window_table(people: Sequence[Person]) -> pd.DataFrame:
    """按任务类型统计需求与时间窗。"""
    rows = []
    for task_type in ("emergency", "production", "shift", "temporary"):
        group = [person for person in people if person.task_type == task_type]
        widths = np.asarray([person.latest - person.earliest for person in group], dtype=float)
        kinds = Counter(demand_kind(person) for person in group)
        service_days = [window_service_days(person) for person in group]
        rows.append({
            "任务类型": task_type,
            "人数": len(group),
            "OD类别数": len({person.od for person in group}),
            "平均时间窗宽度/min": float(np.mean(widths)),
            "最小时间窗宽度/min": float(np.min(widths)),
            "中位时间窗宽度/min": float(np.median(widths)),
            "最大时间窗宽度/min": float(np.max(widths)),
            "出海人数": kinds["出海"],
            "海返人数": kinds["海返"],
            "设施穿梭人数": kinds["设施穿梭"],
            "平均可服务日期数": float(np.mean(service_days)),
            "紧时间窗人数(不超过180min)": int(np.sum(widths <= 180)),
        })
    return pd.DataFrame(rows)


def max_concurrent_aircraft(group: pd.DataFrame) -> int:
    """计算包含30分钟周转的最大同时占用飞机数。"""
    events = []
    for row in group.itertuples(index=False):
        events.append((int(row.start_minute), 1))
        events.append((int(row.turnaround_end_minute), -1))
    active = maximum = 0
    for _, delta in sorted(events, key=lambda item: (item[0], item[1])):
        active += delta
        maximum = max(maximum, active)
    return maximum


def build_daily_fleet_table(flights: pd.DataFrame) -> pd.DataFrame:
    """按日期、机场和机型汇总机队运行。"""
    rows = []
    for day in range(PLANNING_DAYS):
        date = (HORIZON_START + timedelta(days=day)).strftime("%Y-%m-%d")
        for airport in AIRPORTS:
            for aircraft_type in AC:
                group = flights[
                    (flights["day_index"] == day)
                    & (flights["airport"] == airport)
                    & (flights["aircraft_type"] == aircraft_type)
                ]
                aircraft_count = FLEET_COUNTS[airport][aircraft_type]
                task_time = float(group["aircraft_time_min"].sum())
                pass_km = float(group["pass_km"].sum())
                avail_km = float(group["avail_km"].sum())
                rows.append({
                    "日期": date,
                    "机场": airport,
                    "机型": aircraft_type,
                    "飞机数量": aircraft_count,
                    "架次数": len(group),
                    "飞行任务时间/min": task_time,
                    "平均每架飞机任务时间/min": task_time / aircraft_count,
                    "运输人数": int(group["people_count"].sum()),
                    "燃油/kg": float(group["fuel_kg"].sum()),
                    "平均座位利用率": pass_km / avail_km if avail_km > EPS else 0.0,
                    "最早起飞": group["departure_time"].min().strftime("%H:%M") if len(group) else "",
                    "最晚返场": group["return_time"].max().strftime("%H:%M") if len(group) else "",
                    "最大同时占用飞机数": max_concurrent_aircraft(group) if len(group) else 0,
                })
    return pd.DataFrame(rows)


def build_route_structure_table(flights: pd.DataFrame) -> pd.DataFrame:
    """按海上着陆次数统计架次结构。"""
    rows = []
    total = len(flights)
    for landings in range(1, MAX_LANDINGS + 1):
        group = flights[flights["sea_landings"] == landings]
        rows.append({
            "海上着陆次数": landings,
            "架次数": len(group),
            "占比": len(group) / total if total else 0.0,
            "平均运输人数": float(group["people_count"].mean()) if len(group) else 0.0,
            "平均飞机时间/min": float(group["aircraft_time_min"].mean()) if len(group) else 0.0,
            "平均人员在途时间/min": float(group["person_intransit_min"].mean()) if len(group) else 0.0,
            "平均座位利用率": float(group["seat_utilization"].mean()) if len(group) else 0.0,
            "加油架次数": int((group["refuel_count"] > 0).sum()) if len(group) else 0,
            "含设施穿梭任务架次数": int(group["has_shuttle"].sum()) if len(group) else 0,
        })
    return pd.DataFrame(rows)


def write_analysis_workbook(
    args: argparse.Namespace,
    people: Sequence[Person],
    flights: pd.DataFrame,
    tables_dir: Path,
) -> Path:
    """生成包含四张统计表的Excel工作簿。"""
    tables_dir.mkdir(parents=True, exist_ok=True)
    output = tables_dir / "q3_analysis_tables.xlsx"
    tables = {
        "参数设置": build_parameter_table(args),
        "数据与时间窗统计": build_demand_window_table(people),
        "每日机场机队运行统计": build_daily_fleet_table(flights),
        "架次结构统计": build_route_structure_table(flights),
    }
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, frame in tables.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            sheet = writer.sheets[sheet_name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column) + 2
                sheet.column_dimensions[column[0].column_letter].width = min(max(width, 10), 30)
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.0000"
            for header in ("占比", "平均座位利用率"):
                if header in frame.columns:
                    column_index = list(frame.columns).index(header) + 1
                    for row_index in range(2, len(frame) + 2):
                        sheet.cell(row_index, column_index).number_format = "0.00%"
    return output


def plot_two_stage_convergence(summary: dict, figures_dir: Path) -> Path:
    """绘制两阶段最优值和路线池规模。"""
    fig, axes = plt.subplots(2, 1, figsize=(10, 8), constrained_layout=True)
    settings = [
        ("stage1_history", "aircraft_time", "总飞机使用时间/min", "#4C78A8"),
        ("stage2_history", "temporary_served", "临时人员完成数", "#E45756"),
    ]
    for axis, (history_name, metric, ylabel, color) in zip(axes, settings):
        history = summary.get(history_name, [])
        epochs = np.asarray([item.get("epoch", index) for index, item in enumerate(history)])
        values = np.asarray([
            item.get("metrics", {}).get(metric, 0) for item in history
        ], dtype=float)
        pools = np.asarray([item.get("pool_size", 0) for item in history], dtype=float)
        if len(values):
            values = (
                np.minimum.accumulate(values)
                if history_name == "stage1_history"
                else np.maximum.accumulate(values)
            )
        axis.plot(epochs, values, marker="o", linewidth=2, color=color, label=ylabel)
        axis.set_xlabel("epoch")
        axis.set_ylabel(ylabel)
        axis.grid(linestyle=":", alpha=0.35)
        pool_axis = axis.twinx()
        pool_axis.plot(
            epochs, pools, color="#8C8C8C", linestyle="--",
            linewidth=1.4, alpha=0.7, label="路线池规模",
        )
        pool_axis.set_ylabel("路线池规模")
        first_handles, first_labels = axis.get_legend_handles_labels()
        second_handles, second_labels = pool_axis.get_legend_handles_labels()
        axis.legend(
            first_handles + second_handles, first_labels + second_labels,
            loc="best", frameon=False,
        )
    output = figures_dir / "q3_two_stage_convergence.png"
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return output


def plot_aircraft_gantt(flights: pd.DataFrame, figures_dir: Path) -> Path:
    """按日期绘制具体飞机任务和周转时间。"""
    colors = {"T1": "#4C78A8", "T2": "#F58518", "T3": "#54A24B"}
    aircraft_ids = [aircraft_id for aircraft_id, _, _ in FLEET]
    positions = {aircraft_id: index for index, aircraft_id in enumerate(aircraft_ids)}
    fig, axes = plt.subplots(
        PLANNING_DAYS, 1, figsize=(18, 25), sharex=True, sharey=True,
        constrained_layout=True,
    )
    for day, axis in enumerate(np.atleast_1d(axes)):
        group = flights[flights["day_index"] == day]
        for row in group.itertuples(index=False):
            y = positions[row.aircraft_id]
            start_hour = (row.start_minute - day * DAY_MINUTES) / 60.0
            duration_hour = row.aircraft_time_min / 60.0
            axis.barh(
                y, duration_hour, left=start_hour, height=0.65,
                color=colors[row.aircraft_type], edgecolor="white", linewidth=0.4,
            )
            axis.barh(
                y, TURNAROUND / 60.0, left=start_hour + duration_hour,
                height=0.65, color=colors[row.aircraft_type], alpha=0.20,
                edgecolor="none",
            )
        date_label = (HORIZON_START + timedelta(days=day)).strftime("%m-%d")
        axis.text(0.995, 0.93, date_label, transform=axis.transAxes, ha="right", va="top")
        axis.set_yticks(range(len(aircraft_ids)))
        axis.set_yticklabels(aircraft_ids, fontsize=7)
        axis.set_xlim(5.8, 20.6)
        axis.grid(axis="x", linestyle=":", alpha=0.35)
        axis.invert_yaxis()
    axes[-1].set_xlabel("时刻/h")
    axes[-1].set_xticks(range(6, 21, 2))
    legend = [Patch(facecolor=colors[key], label=key) for key in ("T1", "T2", "T3")]
    legend.append(Patch(facecolor="#8C8C8C", alpha=0.20, label="30分钟周转"))
    fig.legend(handles=legend, loc="upper center", ncol=4, frameon=False)
    output = figures_dir / "q3_aircraft_gantt.png"
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return output


def plot_daily_airport_workload(flights: pd.DataFrame, figures_dir: Path) -> Path:
    """绘制各日期、机场和机型的飞机使用时间。"""
    airport_colors = {"A01": "#4C78A8", "A02": "#F58518", "A03": "#54A24B"}
    type_hatches = {"T1": "", "T2": "//", "T3": "xx"}
    dates = [
        (HORIZON_START + timedelta(days=day)).strftime("%m-%d")
        for day in range(PLANNING_DAYS)
    ]
    x = np.arange(PLANNING_DAYS)
    bottom = np.zeros(PLANNING_DAYS)
    fig, axis = plt.subplots(figsize=(11, 6), constrained_layout=True)
    for airport in AIRPORTS:
        for aircraft_type in AC:
            values = np.asarray([
                flights[
                    (flights["day_index"] == day)
                    & (flights["airport"] == airport)
                    & (flights["aircraft_type"] == aircraft_type)
                ]["aircraft_time_min"].sum() / 60.0
                for day in range(PLANNING_DAYS)
            ])
            axis.bar(
                x, values, bottom=bottom, width=0.72,
                color=airport_colors[airport], hatch=type_hatches[aircraft_type],
                edgecolor="white", linewidth=0.5,
            )
            bottom += values
    axis.set_xticks(x)
    axis.set_xticklabels(dates)
    axis.set_xlabel("日期")
    axis.set_ylabel("总飞机使用时间/h")
    axis.grid(axis="y", linestyle=":", alpha=0.35)
    airport_legend = [Patch(facecolor=airport_colors[key], label=key) for key in AIRPORTS]
    type_legend = [
        Patch(facecolor="white", edgecolor="#555555", hatch=type_hatches[key], label=key)
        for key in AC
    ]
    first_legend = axis.legend(handles=airport_legend, loc="upper left", frameon=False)
    axis.add_artist(first_legend)
    axis.legend(handles=type_legend, loc="upper right", frameon=False)
    output = figures_dir / "q3_daily_airport_workload.png"
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return output


def plot_route_structure(flights: pd.DataFrame, figures_dir: Path) -> Path:
    """绘制着陆次数和单架次座位利用率分布。"""
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), constrained_layout=True)
    counts = [int((flights["sea_landings"] == value).sum()) for value in range(1, 6)]
    axes[0].bar(range(1, 6), counts, color="#4C78A8", width=0.7)
    axes[0].set_xlabel("海上着陆次数")
    axes[0].set_ylabel("架次数")
    axes[0].set_xticks(range(1, 6))
    axes[0].grid(axis="y", linestyle=":", alpha=0.35)
    axes[1].hist(
        flights["seat_utilization"].to_numpy(dtype=float) * 100,
        bins=np.linspace(0, 100, 11), color="#F58518", edgecolor="white",
    )
    axes[1].set_xlabel("单架次座位利用率/%")
    axes[1].set_ylabel("架次数")
    axes[1].grid(axis="y", linestyle=":", alpha=0.35)
    output = figures_dir / "q3_route_structure.png"
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return output


def plot_temporary_service_analysis(
    served_assignments: pd.DataFrame,
    people: Sequence[Person],
    figures_dir: Path,
) -> Path:
    """比较已完成和未完成临时人员的时间窗。"""
    served_ids = set(served_assignments["person_id"].astype(str))
    temporary = [person for person in people if not person.mandatory]
    groups = {
        "已完成": [person for person in temporary if person.person_id in served_ids],
        "未完成": [person for person in temporary if person.person_id not in served_ids],
    }
    colors = {"已完成": "#54A24B", "未完成": "#E45756"}
    all_widths = [(person.latest - person.earliest) / 60.0 for person in temporary]
    maximum = max(max(all_widths, default=1.0), 1.0e-6)
    bins = np.linspace(0, maximum, min(16, max(6, len(temporary) // 12 + 1)))
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), constrained_layout=True)
    for label, group in groups.items():
        if not group:
            continue
        widths = [(person.latest - person.earliest) / 60.0 for person in group]
        axes[0].hist(
            widths, bins=bins, alpha=0.55, color=colors[label],
            edgecolor="white", label=f"{label}（{len(group)}人）",
        )
        axes[1].scatter(
            [person.earliest / DAY_MINUTES for person in group],
            [person.latest / DAY_MINUTES for person in group],
            s=25, alpha=0.68, color=colors[label],
            label=f"{label}（{len(group)}人）",
        )
    axes[0].set_xlabel("时间窗宽度/h")
    axes[0].set_ylabel("临时人员数")
    axes[0].grid(axis="y", linestyle=":", alpha=0.35)
    axes[1].set_xlabel("最早可离开时刻/规划日")
    axes[1].set_ylabel("最晚到达时刻/规划日")
    axes[1].grid(linestyle=":", alpha=0.35)
    for axis in axes:
        handles, labels = axis.get_legend_handles_labels()
        if handles:
            axis.legend(frameon=False)
    output = figures_dir / "q3_temporary_service_analysis.png"
    fig.savefig(output, bbox_inches="tight")
    plt.close(fig)
    return output


def generate_reports_from_existing(args: argparse.Namespace) -> Tuple[Path, List[Path]]:
    """只读取现存结果，生成论文表格和图片。"""
    dist_path = report_path(args.dist or args.data_dir / "distances.csv")
    demand_path = report_path(args.demand or args.data_dir / "peopleQ3.csv")
    routes_path = report_path(args.routes_out)
    assignments_path = report_path(args.assign_out)
    summary_path = report_path(args.summary_out)
    for path in (dist_path, demand_path, routes_path, assignments_path, summary_path):
        if not path.exists():
            raise FileNotFoundError(f"报告所需文件不存在：{path}")
    load_distances(dist_path)
    people = load_people(demand_path)
    routes, served_assignments = read_result_frames(routes_path, assignments_path)
    with summary_path.open("r", encoding="utf-8") as stream:
        summary = json.load(stream)
    flights = rebuild_flight_analysis(routes, served_assignments, people)
    verify_rebuilt_metrics(flights, served_assignments, people, summary)
    figures_dir = report_path(args.figures_dir)
    tables_dir = report_path(args.tables_dir)
    figures_dir.mkdir(parents=True, exist_ok=True)
    configure_chinese_plot()
    workbook = write_analysis_workbook(args, people, flights, tables_dir)
    figures = [
        plot_two_stage_convergence(summary, figures_dir),
        plot_aircraft_gantt(flights, figures_dir),
        plot_daily_airport_workload(flights, figures_dir),
        plot_route_structure(flights, figures_dir),
        plot_temporary_service_analysis(served_assignments, people, figures_dir),
    ]
    print(f"analysis tables -> {workbook.resolve()}")
    for figure in figures:
        print(f"analysis figure -> {figure.resolve()}")
    return workbook, figures


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Q3 time-window route-pool and fleet scheduling solver (Python 3.11)"
    )
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--dist", type=Path, default=None)
    parser.add_argument("--demand", type=Path, default=None)
    parser.add_argument("--routes-out", type=Path, default=ROUTES_OUT)
    parser.add_argument("--assign-out", type=Path, default=ASSIGN_OUT)
    parser.add_argument("--summary-out", type=Path, default=SUMMARY_OUT)
    parser.add_argument("--figures-dir", type=Path, default=FIGURES_DIR)
    parser.add_argument("--tables-dir", type=Path, default=TABLES_DIR)
    parser.add_argument(
        "--report-only", action="store_true",
        help="仅读取现存JSON和CSV生成图表，不运行优化搜索",
    )
    parser.add_argument("--time-limit", type=float, default=3600.0)
    parser.add_argument("--stage1-fraction", type=float, default=0.65)
    parser.add_argument("--master-time", type=float, default=120.0)
    parser.add_argument("--epochs", type=int, default=2)
    parser.add_argument("--temp-epochs", type=int, default=5)
    parser.add_argument("--ops-per-epoch", type=int, default=30)
    parser.add_argument("--neighbors", type=int, default=3)
    parser.add_argument("--bootstrap-neighbors", type=int, default=5)
    parser.add_argument("--bootstrap-triples", type=int, default=100)
    parser.add_argument("--temp-candidates-per-route", type=int, default=2)
    parser.add_argument("--route-variants", type=int, default=4)
    parser.add_argument("--max-patterns", type=int, default=10000)
    parser.add_argument(
        "--stage1-pool-fraction", type=float, default=0.55,
        help="阶段一路线池软上限占总上限的比例",
    )
    parser.add_argument(
        "--stage2-bootstrap-pool-fraction", type=float, default=0.80,
        help="阶段二首次求解前路线池软上限占总上限的比例",
    )
    parser.add_argument("--max-route-people", type=int, default=38)
    parser.add_argument("--assignment-limit", type=int, default=300)
    parser.add_argument("--workers", type=int, default=min(4, os.cpu_count() or 1))
    parser.add_argument("--seed", type=int, default=2026)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.report_only:
        generate_reports_from_existing(args)
        return
    started = time.monotonic()
    total_deadline = started + max(1.0, args.time_limit)
    stage1_fraction = min(0.9, max(0.2, args.stage1_fraction))
    stage1_deadline = started + args.time_limit * stage1_fraction
    rng = random.Random(args.seed)
    dist_path = (args.dist or args.data_dir / "distances.csv").resolve()
    demand_path = (args.demand or args.data_dir / "peopleQ3.csv").resolve()
    print(f"读取距离：{dist_path}")
    print(f"读取需求：{demand_path}")
    load_distances(dist_path)
    people = load_people(demand_path)
    mandatory_ids = [person.index for person in people if person.mandatory]
    temporary_ids = [person.index for person in people if not person.mandatory]
    type_counts = Counter(person.task_type for person in people)
    print(
        f"people={len(people)}, mandatory={len(mandatory_ids)}, temporary={len(temporary_ids)}, "
        f"emergency={type_counts['emergency']}, production={type_counts['production']}, "
        f"shift={type_counts['shift']}, fleet={len(FLEET)}"
    )
    factory = RouteFactory(
        people,
        max_route_people=max(MAX_SEATS, args.max_route_people),
        assignment_limit=max(1, args.assignment_limit),
    )
    pool = RoutePool(max(100, args.max_patterns))
    variants = max(1, args.route_variants)
    stage1_pool_fraction = min(0.85, max(0.25, args.stage1_pool_fraction))
    stage2_bootstrap_pool_fraction = min(
        0.95,
        max(stage1_pool_fraction + 0.10, args.stage2_bootstrap_pool_fraction),
    )

    print("\n=== 阶段一：只安排应急、生产和常规倒班任务 ===")
    print("生成时间窗可行的初始批次……")
    mandatory_batches = build_initial_pool(
        people, mandatory_ids, factory, pool, variants,
        "mandatory_initial", add_splits=False,
    )
    stage1_pool_limit = pool.set_active_limit(max(
        len(pool), math.floor(pool.max_patterns * stage1_pool_fraction),
    ))
    print(
        f"阶段一路线池配额：{stage1_pool_limit}/{pool.max_patterns} "
        f"（为阶段二保留{pool.max_patterns - stage1_pool_limit}）"
    )
    print("首次排班前生成跨OD联合运输候选……")
    bootstrap_added = bootstrap_merge_batches(
        mandatory_batches, people, pool, factory, variants,
        max(1, args.bootstrap_neighbors), max(0, args.bootstrap_triples),
        rng, stage1_deadline,
    )
    print(f"  bootstrap新增候选={bootstrap_added}，路线池={len(pool)}")
    remaining = max(0.1, min(args.master_time, stage1_deadline - time.monotonic()))
    first = solve_master(
        pool, people, mandatory_ids, temporary_ids,
        stage=1, baseline_cap=None, incumbent=None,
        time_limit=remaining, workers=args.workers, seed=args.seed,
    )
    current = first
    stage1_history = [{
        "epoch": 0, "pool_size": len(pool), "metrics": current.metrics.__dict__,
        "pool_optimal": current.pool_optimal, "stages": current.stages,
    }]
    print_metrics("阶段一初始排班：", current.metrics, len(pool))

    stage1_stop_reason = "configured_epochs_completed"
    for epoch in range(1, max(0, args.epochs) + 1):
        if time.monotonic() >= stage1_deadline:
            stage1_stop_reason = "stage1_time_limit"
            break
        if pool.is_full:
            stage1_stop_reason = "stage1_pool_soft_limit"
            break
        added = expand_selected_pool(
            current.selections, pool, factory, variants,
            max(1, args.neighbors), max(0, args.ops_per_epoch), rng,
            stage1_deadline, f"mandatory_epoch_{epoch}",
        )
        if added == 0:
            stage1_stop_reason = "no_new_patterns"
            print(f"阶段一 epoch {epoch}: 没有新增候选，停止扩充。")
            break
        remaining = min(args.master_time, stage1_deadline - time.monotonic())
        if remaining <= 0.05:
            stage1_stop_reason = "stage1_time_limit"
            break
        result = solve_master(
            pool, people, mandatory_ids, temporary_ids,
            stage=1, baseline_cap=None, incumbent=current.selections,
            time_limit=remaining, workers=args.workers, seed=args.seed + epoch,
        )
        improved = stage1_better(result.metrics, current.metrics)
        if improved:
            current = result
        stage1_history.append({
            "epoch": epoch, "pool_size": len(pool), "new_patterns": added,
            "accepted": improved, "metrics": result.metrics.__dict__,
            "pool_optimal": result.pool_optimal, "stages": result.stages,
        })
        print_metrics(
            f"阶段一 epoch {epoch}（新增{added}，accepted={improved}）：",
            current.metrics, len(pool),
        )

    baseline_time = current.metrics.aircraft_time
    mandatory_solution = current
    validate_solution(current.selections, pool, people, None)
    print(f"\n阶段一时间基准 T0 = {baseline_time} min")

    print("\n=== 阶段二：在 T<=T0 下最大化临时任务人数 ===")
    print("生成临时任务单独架次、批次及搭载候选……")
    stage2_bootstrap_pool_limit = pool.set_active_limit(max(
        len(pool), math.floor(pool.max_patterns * stage2_bootstrap_pool_fraction),
    ))
    print(
        f"阶段二初始路线池配额：{stage2_bootstrap_pool_limit}/{pool.max_patterns} "
        f"（为后续temp-epochs保留{pool.max_patterns - stage2_bootstrap_pool_limit}）"
    )
    for person_id in temporary_ids:
        add_generated(pool, factory, (person_id,), min(2, variants), "temporary_single")
        if pool.is_full:
            break
    if not pool.is_full:
        build_initial_pool(
            people, temporary_ids, factory, pool, min(2, variants),
            "temporary_batch", add_splits=False,
        )
    expand_temporary_pool(
        mandatory_solution.selections, temporary_ids, pool, factory, variants,
        max(1, args.temp_candidates_per_route), total_deadline,
        "temporary_bootstrap",
    )
    remaining = min(args.master_time, max(0.1, total_deadline - time.monotonic()))
    stage2_result = solve_master(
        pool, people, mandatory_ids, temporary_ids,
        stage=2, baseline_cap=baseline_time,
        incumbent=mandatory_solution.selections,
        time_limit=remaining, workers=args.workers, seed=args.seed + 1000,
    )
    current2 = stage2_result if stage2_better(
        stage2_result.metrics, mandatory_solution.metrics
    ) else MasterResult(
        list(mandatory_solution.selections), mandatory_solution.metrics,
        stage2_result.stages, False,
    )
    stage2_history = [{
        "epoch": 0, "pool_size": len(pool), "metrics": current2.metrics.__dict__,
        "pool_optimal": stage2_result.pool_optimal, "stages": stage2_result.stages,
    }]
    print_metrics("阶段二初始结果：", current2.metrics, len(pool))

    # 初始临时候选求解完成后，才开放最后一段硬容量给temp-epochs。
    pool.set_active_limit(pool.max_patterns)
    print(
        f"阶段二LNS路线池已开放至硬上限：{pool.max_patterns}，"
        f"当前剩余容量={pool.remaining_capacity}"
    )

    stage2_stop_reason = "configured_epochs_completed"
    for epoch in range(1, max(0, args.temp_epochs) + 1):
        if time.monotonic() >= total_deadline:
            stage2_stop_reason = "total_time_limit"
            break
        if pool.is_full:
            stage2_stop_reason = "hard_pool_limit"
            break
        added_temp = expand_temporary_pool(
            current2.selections, temporary_ids, pool, factory, variants,
            max(1, args.temp_candidates_per_route), total_deadline,
            f"temporary_epoch_{epoch}",
        )
        added_general = expand_selected_pool(
            current2.selections, pool, factory, variants,
            max(1, args.neighbors // 2), max(0, args.ops_per_epoch // 2), rng,
            total_deadline, f"stage2_epoch_{epoch}",
        )
        added = added_temp + added_general
        if added == 0:
            stage2_stop_reason = "no_new_patterns"
            print(f"阶段二 epoch {epoch}: 没有新增候选，停止扩充。")
            break
        remaining = min(args.master_time, total_deadline - time.monotonic())
        if remaining <= 0.05:
            stage2_stop_reason = "total_time_limit"
            break
        result = solve_master(
            pool, people, mandatory_ids, temporary_ids,
            stage=2, baseline_cap=baseline_time,
            incumbent=current2.selections,
            time_limit=remaining, workers=args.workers,
            seed=args.seed + 1000 + epoch,
        )
        improved = stage2_better(result.metrics, current2.metrics)
        if improved:
            current2 = result
        stage2_history.append({
            "epoch": epoch, "pool_size": len(pool), "new_patterns": added,
            "accepted": improved, "metrics": result.metrics.__dict__,
            "pool_optimal": result.pool_optimal, "stages": result.stages,
        })
        print_metrics(
            f"阶段二 epoch {epoch}（新增{added}，accepted={improved}）：",
            current2.metrics, len(pool),
        )

    final_metrics = validate_solution(
        current2.selections, pool, people, baseline_time,
    )
    write_outputs(
        current2.selections, pool, people,
        args.routes_out, args.assign_out,
    )
    elapsed_s = round(time.monotonic() - started, 3)
    summary = {
        "algorithm": "Q3 time-window route pool + fleet interval scheduling + LNS",
        "python": sys.version.split()[0],
        "people": len(people),
        "mandatory_people": len(mandatory_ids),
        "temporary_people": len(temporary_ids),
        "fleet_size": len(FLEET),
        "stage1_baseline_aircraft_time_min": baseline_time,
        "final_metrics": {
            "temporary_served": final_metrics.temporary_served,
            "temporary_service_rate": (
                final_metrics.temporary_served / len(temporary_ids)
                if temporary_ids else 1.0
            ),
            "total_aircraft_time_min": final_metrics.aircraft_time,
            "total_person_intransit_min": final_metrics.person_intransit,
            "sorties": final_metrics.sorties,
            "total_fuel_kg": round(final_metrics.fuel_kg, 6),
            "seat_utilization": final_metrics.seat_utilization,
        },
        "objective_order_stage1": [
            "total_aircraft_time_min", "total_person_intransit_min",
            "sorties", "total_fuel_kg", "seat_utilization_max",
        ],
        "objective_order_stage2": [
            "temporary_served_max", "total_aircraft_time_min",
            "total_person_intransit_min", "sorties",
            "total_fuel_kg", "seat_utilization_max",
        ],
        "route_pool_size": len(pool),
        "route_pool_limits": {
            "hard_limit": pool.max_patterns,
            "stage1_soft_limit": stage1_pool_limit,
            "stage2_bootstrap_soft_limit": stage2_bootstrap_pool_limit,
            "final_remaining_capacity": pool.max_patterns - len(pool),
        },
        "stage1_stop_reason": stage1_stop_reason,
        "stage2_stop_reason": stage2_stop_reason,
        "route_factory_calls": factory.calls,
        "route_factory_cache_hits": factory.cache_hits,
        "stage1_history": stage1_history,
        "stage2_history": stage2_history,
        "claim_scope": "best valid solution found; exact claims apply only to solved stages over generated pool",
        "original_problem_global_optimum_claimed": False,
        "configured_time_limit_s": args.time_limit,
        "elapsed_s": elapsed_s,
        "unused_time_budget_s": round(max(0.0, args.time_limit - elapsed_s), 3),
        "routes_out": str(args.routes_out.resolve()),
        "assignments_out": str(args.assign_out.resolve()),
    }
    args.summary_out.parent.mkdir(parents=True, exist_ok=True)
    with args.summary_out.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    generate_reports_from_existing(args)
    print_metrics("\n=== Q3 最终结果 ===", final_metrics, len(pool))
    print(f"临时任务满足率：{final_metrics.temporary_served}/{len(temporary_ids)}")
    print("时间窗、具体飞机、周转、取送、容量、燃油与人员覆盖完整校验通过。")
    print(f"routes -> {args.routes_out.resolve()}")
    print(f"assignments -> {args.assign_out.resolve()}")
    print(f"summary -> {args.summary_out.resolve()}")
    print(f"阶段一停止原因：{stage1_stop_reason}")
    print(f"阶段二停止原因：{stage2_stop_reason}")
    if stage2_stop_reason == "hard_pool_limit" and elapsed_s + 1 < args.time_limit:
        print(
            f"注意：因路线池达到硬上限而提前结束，尚余约"
            f"{args.time_limit - elapsed_s:.0f}秒未使用；如需继续搜索，可适度增大"
            f"--max-patterns，或减小候选生成参数。"
        )
    print("最优性声明仅限已生成候选池中、且求解器返回OPTIMAL的阶段。")


if __name__ == "__main__":
    main()
